"""
Build an Excel statement deterministically from an NL spec + parsed bills.
Numbers come straight from the extracted rows; the spec only says what to pull.
"""
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL = "1A7F77"
WHITE = "FFFFFF"
_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _bill_key(bn):
    m = re.search(r"(\d+)", bn or "")
    return int(m.group(1)) if m else 9999


def _hdr(ws, r, c, t):
    cell = ws.cell(r, c, t)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def build_from_spec(parsed, spec):
    """spec: {value_column, include_total, selections:[{schedule,item}]}"""
    value_col = spec.get("value_column") or "Amount Since last Bill including special condition"
    include_total = bool(spec.get("include_total"))
    sel = spec.get("selections") or []
    sel_set = {(s.get("schedule"), s.get("item")) for s in sel}

    bills = sorted(parsed, key=lambda b: _bill_key(b.get("bill_no")))
    bill_nos = [b.get("bill_no") or f"B{i+1}" for i, b in enumerate(bills)]

    # amount[(sched,item)][bill_no] = value
    amounts = {}
    descriptions = {}
    for b in bills:
        bn = b.get("bill_no")
        for r in b["rows"]:
            key = (r["schedule"], r["item"])
            if key in sel_set:
                amounts.setdefault(key, {})[bn] = r["values"].get(value_col) or 0
                descriptions.setdefault(key, r["description"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.cell(1, 1, "STATEMENT").font = Font(bold=True, size=12)
    ws.cell(2, 1, f"Value column: {value_col}").font = Font(size=9, italic=True)

    headers = ["Schedule", "Item No", "Description"] + bill_nos
    if include_total:
        headers.append("Total")
    for c, h in enumerate(headers, 1):
        _hdr(ws, 4, c, h)

    r = 5
    col_totals = {bn: 0.0 for bn in bill_nos}
    grand = 0.0
    for key in sorted(amounts.keys()):
        sched, item = key
        ws.cell(r, 1, sched).border = BORDER
        ws.cell(r, 2, item).border = BORDER
        dc = ws.cell(r, 3, (descriptions.get(key) or "")[:80]); dc.border = BORDER
        row_total = 0.0
        for i, bn in enumerate(bill_nos):
            v = amounts[key].get(bn, 0) or 0
            cell = ws.cell(r, 4 + i, round(v, 2)); cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
            col_totals[bn] += v
            row_total += v
        if include_total:
            tc = ws.cell(r, 4 + len(bill_nos), round(row_total, 2))
            tc.border = BORDER; tc.font = Font(bold=True); tc.alignment = Alignment(horizontal="right")
        grand += row_total
        r += 1

    # totals row
    tr = ws.cell(r, 1, "TOTAL"); tr.font = Font(bold=True, color=WHITE); tr.fill = PatternFill("solid", fgColor=TEAL); tr.border = BORDER
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=TEAL); ws.cell(r, 2).border = BORDER
    ws.cell(r, 3).fill = PatternFill("solid", fgColor=TEAL); ws.cell(r, 3).border = BORDER
    for i, bn in enumerate(bill_nos):
        cell = ws.cell(r, 4 + i, round(col_totals[bn], 2))
        cell.font = Font(bold=True, color=WHITE); cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.border = BORDER; cell.alignment = Alignment(horizontal="right")
    if include_total:
        cell = ws.cell(r, 4 + len(bill_nos), round(grand, 2))
        cell.font = Font(bold=True, color=WHITE); cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.border = BORDER; cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 44
    for i in range(len(bill_nos) + (1 if include_total else 0)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 13
    ws.freeze_panes = "D5"

    info = {"rows": len(amounts), "value_column": value_col,
            "include_total": include_total, "bills": bill_nos,
            "schedules": sorted({k[0] for k in amounts})}
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, info
