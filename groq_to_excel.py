"""
Convert Groq-extracted JSON to Excel based on the prompt template output format.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL = "1A7F77"
TEAL_DARK = "0E5C55"
WHITE = "FFFFFF"
_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(ws, r, c, text):
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=TEAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def _num(ws, r, c, v):
    cell = ws.cell(r, c, round(v, 2) if isinstance(v, (int, float)) else v)
    cell.border = BORDER
    if isinstance(v, (int, float)):
        cell.alignment = Alignment(horizontal="right")
    return cell


def groq_json_to_excel(groq_data, template):
    """
    Convert Groq JSON to Excel based on the template's output_format.
    
    groq_data: parsed JSON from Groq
    template: dict with keys 'output_format', 'columns', etc.
    
    Returns: BytesIO buffer with .xlsx
    """
    output_fmt = template.get("output_format", "Statement")
    cols = template.get("columns", [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = output_fmt[:31]
    
    # Header
    for c, h in enumerate(cols, 1):
        _hdr(ws, 1, c, h)
    
    # Data rows
    r = 2
    
    if "bills" in groq_data:
        # Bill summary format
        for item in groq_data.get("bills", []):
            ws.cell(r, 1, item.get("bill_no", "")).border = BORDER
            ws.cell(r, 2, item.get("meas_start", "")).border = BORDER
            ws.cell(r, 3, item.get("meas_end", "")).border = BORDER
            _num(ws, r, 4, item.get("amount", 0))
            r += 1
    
    elif "items" in groq_data:
        # Item statement (single or multi-item, with or without totals)
        items = groq_data.get("items", [])
        
        # Group by (schedule, item_no) for multi-bill format
        grouped = {}
        for item in items:
            key = (item.get("schedule"), item.get("item_no"))
            if key not in grouped:
                grouped[key] = {"schedule": key[0], "item_no": key[1], "bills": {}}
            bill = item.get("bill_no", "")
            if bill:
                grouped[key]["bills"][bill] = item.get("amount", 0)
        
        # Write rows
        for (sch, itemno), data in sorted(grouped.items()):
            ws.cell(r, 1, f"{sch}-{itemno}").border = BORDER
            
            # Bills columns (B1-B10)
            total = 0.0
            for i, col_header in enumerate(cols[1:], start=2):  # Skip the "Sch and Item No" column
                if col_header.startswith("B"):
                    amt = data["bills"].get(col_header, 0)
                    _num(ws, r, i, amt)
                    total += amt
                elif col_header == "Total":
                    _num(ws, r, i, total)
            r += 1
    
    # Widths
    for c, w in enumerate([16, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13], 1):
        if c <= len(cols):
            ws.column_dimensions[get_column_letter(c)].width = w
    
    ws.freeze_panes = "A2"
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
