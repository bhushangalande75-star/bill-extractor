"""
Build the bill-wise statement workbook from parsed bills + a filter config.

Output mirrors the railway "STATEMENT SHOWING BILL WISE ... CONSUMPTION" format:
rows grouped by schedule, one column per bill, plus a Total column and a
grand TOTAL row.
"""
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEAL = "14B8A6"
TEAL_DARK = "0F766E"
LIGHT = "E1F5EE"
WHITE = "FFFFFF"

_thin = Side(style="thin", color="C9D6D1")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _bill_sort_key(b):
    m = re.search(r"\d+", b.get("bill_no") or "")
    return int(m.group()) if m else 999


def build_workbook(parsed_bills, config):
    """
    parsed_bills: list of dicts from extractor.extract_rows()
    config: {
        "value_column": str,
        "items":     [str, ...],          # item numbers to include
        "schedules": [str, ...] or [],     # empty => all schedules present
        "title":     str (optional),
        "ca_no":     str (optional),
    }
    Returns an in-memory .xlsx (BytesIO).
    """
    value_col = config["value_column"]
    want_items = list(dict.fromkeys(config["items"]))          # preserve order, dedupe
    want_scheds = set(config.get("schedules") or [])

    bills = sorted(parsed_bills, key=_bill_sort_key)
    bill_labels = [b.get("bill_no") or f"B{i+1}" for i, b in enumerate(bills)]

    # lookup[(schedule, item)][bill_label] = value
    lookup = {}
    scheds_present = set()
    for b, label in zip(bills, bill_labels):
        for row in b["rows"]:
            scheds_present.add(row["schedule"])
            key = (row["schedule"], row["item"])
            val = row["values"].get(value_col)
            if val is not None:
                lookup.setdefault(key, {})
                lookup[key][label] = lookup[key].get(label, 0.0) + val

    scheds = sorted(s for s in scheds_present if s and (not want_scheds or s in want_scheds))

    wb = Workbook()
    ws = wb.active
    ws.title = "Bill-wise Statement"

    ncols = 2 + len(bill_labels) + 1   # Sch&Item | bills... | Total
    last_col = ncols

    # ---- title block ----
    ws.cell(1, 1, config.get("title") or "STATEMENT SHOWING BILL-WISE AMOUNT")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1).font = Font(bold=True, size=13, color=TEAL_DARK)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    ws.cell(2, 1, f"CA No.: {config.get('ca_no') or '-'}    |    Value: {value_col}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1).font = Font(italic=True, size=10, color="555555")
    ws.cell(2, 1).alignment = Alignment(horizontal="center")

    # ---- header row ----
    hdr = 4
    headers = ["SN", "Sch & Item No"] + bill_labels + ["Total"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(hdr, c, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    r = hdr + 1
    sn = 0
    col_totals = {label: 0.0 for label in bill_labels}
    grand_total = 0.0

    for sched in scheds:
        items_here = [it for it in want_items if (sched, it) in lookup]
        if not items_here:
            continue  # no selected item in this schedule -> no header

        # schedule sub-header
        scell = ws.cell(r, 2, f"SCH-{sched}")
        scell.font = Font(bold=True, color=TEAL_DARK)
        for c in range(1, last_col + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
            ws.cell(r, c).border = BORDER
        r += 1

        for item in items_here:
            key = (sched, item)
            sn += 1
            ws.cell(r, 1, sn).border = BORDER
            ws.cell(r, 2, item).border = BORDER
            row_total = 0.0
            for ci, label in enumerate(bill_labels, start=3):
                v = lookup[key].get(label)
                cell = ws.cell(r, ci)
                if v is not None:
                    cell.value = round(v, 2)
                    row_total += v
                    col_totals[label] += v
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="right")
            tcell = ws.cell(r, last_col, round(row_total, 2))
            tcell.border = BORDER
            tcell.font = Font(bold=True)
            tcell.alignment = Alignment(horizontal="right")
            grand_total += row_total
            r += 1

    # ---- grand total row ----
    ws.cell(r, 2, "TOTAL").font = Font(bold=True, color=WHITE)
    for c in range(1, last_col + 1):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=TEAL_DARK)
        ws.cell(r, c).border = BORDER
    for ci, label in enumerate(bill_labels, start=3):
        cell = ws.cell(r, ci, round(col_totals[label], 2))
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="right")
    gcell = ws.cell(r, last_col, round(grand_total, 2))
    gcell.font = Font(bold=True, color=WHITE)
    gcell.alignment = Alignment(horizontal="right")

    # ---- widths ----
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    from openpyxl.utils import get_column_letter
    for c in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "C5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ----------------------------------------------------------------------------
# Multi-sheet preset workbooks
# ----------------------------------------------------------------------------
from openpyxl.utils import get_column_letter


def _hdr_cell(ws, r, c, text, fill=TEAL, color=WHITE):
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    return cell


def _build_lookup(parsed_bills, value_col):
    """Return (lookup, present_pairs, bill_labels)."""
    bills = sorted(parsed_bills, key=_bill_sort_key)
    labels = [b.get("bill_no") or f"B{i+1}" for i, b in enumerate(bills)]
    lookup, present = {}, set()
    for b, label in zip(bills, labels):
        for row in b["rows"]:
            key = (row["schedule"], row["item"])
            present.add(key)
            v = row["values"].get(value_col)
            if v is not None:
                lookup.setdefault(key, {})
                lookup[key][label] = lookup[key].get(label, 0.0) + v
    return lookup, present, labels


def _row_plan(select, present):
    """Return ordered list of (schedule, item) pairs to print."""
    pairs = []
    if select["mode"] == "items":
        items = list(dict.fromkeys(select["items"]))
        scheds = sorted({s for (s, _it) in present})
        for s in scheds:
            for it in items:
                if (s, it) in present:
                    pairs.append((s, it))
    else:  # map: explicit schedule -> items, always shown
        for s, items in select["map"].items():
            for it in items:
                pairs.append((s, it))
    return pairs


def _write_statement_sheet(ws, spec, lookup, labels, value_col):
    layout = spec.get("layout", "combined")
    total = spec.get("total", False)
    pairs = _row_plan(spec["select"], set(lookup) | _all_present)

    # header
    ws.cell(1, 1, f"Value: {value_col}").font = Font(italic=True, size=9, color="555555")
    if layout == "separate":
        headers = ["SN", "Sch", "Item No"] + labels + (["Total"] if total else [])
        key_cols = 3
    else:
        headers = ["SN", "Sch & Item No"] + labels + (["Total"] if total else [])
        key_cols = 2
    hr = 2
    for c, h in enumerate(headers, start=1):
        _hdr_cell(ws, hr, c, h)

    r = hr + 1
    col_tot = {lab: 0.0 for lab in labels}
    sn = 0
    for (s, it) in pairs:
        sn += 1
        ws.cell(r, 1, sn).border = BORDER
        if layout == "separate":
            ws.cell(r, 2, s).border = BORDER
            ws.cell(r, 3, it).border = BORDER
        else:
            ws.cell(r, 2, f"{s}-{it}").border = BORDER
        row_tot = 0.0
        first_bill_col = key_cols + 1
        for i, lab in enumerate(labels):
            v = lookup.get((s, it), {}).get(lab)
            v = 0.0 if v is None else v
            cell = ws.cell(r, first_bill_col + i, round(v, 2))
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
            row_tot += v
            col_tot[lab] += v
        if total:
            tc = ws.cell(r, first_bill_col + len(labels), round(row_tot, 2))
            tc.border = BORDER
            tc.font = Font(bold=True)
            tc.alignment = Alignment(horizontal="right")
        r += 1

    # total row
    _hdr_cell(ws, r, key_cols, "TOTAL", fill=TEAL_DARK)
    for c in range(1, key_cols):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=TEAL_DARK)
        ws.cell(r, c).border = BORDER
    grand = 0.0
    fb = key_cols + 1
    for i, lab in enumerate(labels):
        cell = ws.cell(r, fb + i, round(col_tot[lab], 2))
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL_DARK)
        cell.alignment = Alignment(horizontal="right")
        grand += col_tot[lab]
    if total:
        gc = ws.cell(r, fb + len(labels), round(grand, 2))
        gc.font = Font(bold=True, color=WHITE)
        gc.fill = PatternFill("solid", fgColor=TEAL_DARK)
        gc.alignment = Alignment(horizontal="right")

    # widths
    ws.column_dimensions["A"].width = 5
    for c in range(2, key_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12 if layout == "separate" else 16
    end = key_cols + len(labels) + (1 if total else 0)
    for c in range(key_cols + 1, end + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = ws.cell(hr + 1, key_cols + 1).coordinate


def _write_bill_summary(ws, parsed_bills):
    bills = sorted(parsed_bills, key=_bill_sort_key)
    headers = ["Bill No", "Measurement Start", "Measurement Complete", "Bill Amount (incl GST)"]
    for c, h in enumerate(headers, start=1):
        _hdr_cell(ws, 1, c, h)
    r = 2
    total = 0.0
    for b in bills:
        ws.cell(r, 1, b.get("bill_no") or "").border = BORDER
        ws.cell(r, 2, b.get("meas_from") or "").border = BORDER
        ws.cell(r, 3, b.get("meas_to") or "").border = BORDER
        amt = b.get("bill_amount")
        cell = ws.cell(r, 4, round(amt, 2) if amt is not None else "")
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right")
        if amt:
            total += amt
        r += 1
    _hdr_cell(ws, r, 1, "TOTAL", fill=TEAL_DARK)
    for c in (2, 3):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=TEAL_DARK)
        ws.cell(r, c).border = BORDER
    tc = ws.cell(r, 4, round(total, 2))
    tc.font = Font(bold=True, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=TEAL_DARK)
    tc.alignment = Alignment(horizontal="right")
    for col, w in zip("ABCD", (10, 20, 20, 22)):
        ws.column_dimensions[col].width = w


# module-level holder so _write_statement_sheet can see all present pairs
_all_present = set()


def build_preset_workbook(parsed_bills, preset):
    """Build a multi-sheet workbook from a preset spec."""
    global _all_present
    wb = Workbook()
    wb.remove(wb.active)
    default_val = "Amount Since last Bill including special condition"

    for spec in preset["sheets"]:
        title = spec["title"][:31]  # Excel sheet-name limit
        ws = wb.create_sheet(title=title)
        if spec["type"] == "bill_summary":
            _write_bill_summary(ws, parsed_bills)
        else:
            value_col = spec.get("value_column") or default_val
            lookup, present, labels = _build_lookup(parsed_bills, value_col)
            _all_present = present
            _write_statement_sheet(ws, spec, lookup, labels, value_col)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ----------------------------------------------------------------------------
# Raw "as-is" full extract — every column and row of the bill, one sheet per bill
# ----------------------------------------------------------------------------
RAW_HEADERS = ["Schedule", "Sr", "Item No", "Unit", "Description",
               "Base Rate", "Agmt Rate", "Orig Agmt Qty", "Curr Agmt Qty",
               "Qty upto last", "Qty since last", "Qty upto date",
               "Amt upto last", "Amt since last", "Amt incl spec",
               "Total upto date", "Remarks"]
RAW_WIDTHS = [10, 5, 11, 8, 44, 11, 11, 12, 12, 12, 12, 12, 13, 13, 14, 14, 16]


def build_raw_workbook(parsed_bills):
    """Faithful dump: one sheet per bill with every item row and every column."""
    wb = Workbook()
    wb.remove(wb.active)

    ws0 = wb.create_sheet("Bills")
    _write_bill_summary(ws0, parsed_bills)

    used = {"Bills"}
    for b in sorted(parsed_bills, key=_bill_sort_key):
        base = (b.get("bill_no") or "Bill")[:28]
        name, i = base, 2
        while name in used:
            name = f"{base}_{i}"
            i += 1
        used.add(name)
        ws = wb.create_sheet(name)

        info = (f"{b.get('ca_no') or ''}   |   Bill {b.get('bill_no') or ''}   |   "
                f"Meas {b.get('meas_from') or '-'} to {b.get('meas_to') or '-'}   |   "
                f"Bill amount {b.get('bill_amount') if b.get('bill_amount') is not None else '-'}")
        ws.cell(1, 1, info).font = Font(italic=True, size=10, color="555555")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RAW_HEADERS))

        for c, h in enumerate(RAW_HEADERS, start=1):
            _hdr_cell(ws, 3, c, h)

        r = 4
        for row in b["rows"]:
            f = row["full"]
            vals = [row["schedule"], f["sr_no"], row["item"], row["unit"], row["description"],
                    f["base_rate"], f["agmt_rate"], f["orig_qty"], f["curr_qty"],
                    f["qty_upto_last"], f["qty_since_last"], f["qty_upto_date"],
                    f["amt_upto_last"], f["amt_since_last"], f["amt_incl_spec"],
                    f["total_upto_date"], f["remarks"]]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c, v if v is not None else "")
                cell.border = BORDER
                if c >= 6 and isinstance(v, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
            r += 1

        for c, w in enumerate(RAW_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
