"""
Assemble the GCC 46A PVC workbook in the Manoj-Sethi layout from:
  - parsed index data banks (index_bank.load_all)
  - a contract config: base months per agreement + per-quarter gross breakdown

Produces the headline 'Summary' sheet (two blocks 9A/9B/9C and 9D + G.Total)
and a 'GPVC' sheet.  Numbers come from the validated engine in pvc.py.
"""
import io
import pvc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL = "1A7F77"
TEAL_DARK = "0E5C55"
LIGHT = "E8F4F2"
WHITE = "FFFFFF"
_thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

BAL_W = {"labour": 0.20, "machinery": 0.30, "diesel": 0.15, "material": 0.20}
OTH_W = {"labour": 0.10, "steel_other": 0.50, "machinery": 0.10, "diesel": 0.10, "material": 0.05}
REINF_W = {"tmt": 0.85}


def _hdr(ws, r, c, text, fill=TEAL, color=WHITE, bold=True):
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=bold, color=color, size=9)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def _num(ws, r, c, v, bold=False, fill=None):
    cell = ws.cell(r, c, round(v, 2) if isinstance(v, (int, float)) else v)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="right")
    if bold:
        cell.font = Font(bold=True, color=WHITE if fill else "000000", size=9)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def _parts(gross, weights, months, base, indices, snap=()):
    """Per-component PVC for one entry, rounded to 2dp like the workbook."""
    out = {}
    for comp, w in weights.items():
        s = indices.get(comp, {})
        bv = s.get(base)
        ms = months[:1] if comp in snap else months
        av = pvc.avg_index(s, ms)
        if av is not None:
            av = round(av, 2)
        out[comp] = round(pvc.pvc_one(gross, w, bv, av), 2)
    return out


def build_pvc_workbook(indices, config):
    """config: {
        'name','ca_no','agency','base_month',
        'base_months': {'Original':'YYYY-MM', 'SA-1':..., ...},
        'entries': [ {sn, qtr, agreement, months, snap(list),
                      balance, reinf, other, cement, plate}, ... ]
    }"""
    wb = Workbook()
    bm = config.get("base_months", {})

    # ---------------- Summary sheet ----------------
    ws = wb.active
    ws.title = "Summary"
    ws.cell(1, 1, "SUMMARY OF PRICE VARIATION BILL").font = Font(bold=True, size=12)
    ws.cell(2, 1, f"Name of work: {config.get('name','')}").font = Font(size=9)
    ws.cell(3, 1, f"CA No.: {config.get('ca_no','')}    Agency: {config.get('agency','')}"
                  f"    Base Month: {config.get('base_month','')}").font = Font(size=9)

    cols_I = ["SN", "Qtr", "Labour", "Plant\nMachinery", "Fuel &\nLubricants",
              "Other\nMaterial", "Cement", "Reinforcement\nBars", "Angles,\nchannel",
              "Plates,\nFlats", "TOTAL"]

    def write_block(start_row, title, rows, kind):
        _hdr(ws, start_row, 1, title, fill=TEAL_DARK)
        for c in range(2, len(cols_I) + 1):
            ws.cell(start_row, c).fill = PatternFill("solid", fgColor=TEAL_DARK)
            ws.cell(start_row, c).border = BORDER
        hr = start_row + 1
        for c, h in enumerate(cols_I, 1):
            _hdr(ws, hr, c, h)
        r = hr + 1
        tot = [0.0] * 9   # labour,pm,fuel,material,cement,reinf,steel,plate,TOTAL
        for e in rows:
            base = bm.get(e["agreement"], config.get("base_month_key"))
            snap = e.get("snap", [])
            if kind == "I":
                bp = _parts(e.get("balance", 0), BAL_W, e["months"], base, indices, snap)
                rp = _parts(e.get("reinf", 0), REINF_W, e["months"], base, indices, snap)
                vals = [bp["labour"], bp["machinery"], bp["diesel"], bp["material"],
                        0.0, rp["tmt"], 0.0, 0.0]
            else:
                op = _parts(e.get("other", 0), OTH_W, e["months"], base, indices, snap)
                vals = [op["labour"], op["machinery"], op["diesel"], op["material"],
                        0.0, 0.0, op["steel_other"], 0.0]
            row_total = round(sum(vals), 2)
            ws.cell(r, 1, e["sn"]).border = BORDER
            ws.cell(r, 2, e["qtr"]).border = BORDER
            for i, v in enumerate(vals):
                _num(ws, r, 3 + i, v)
                tot[i] += v
            _num(ws, r, 11, row_total, bold=True)
            tot[8] += row_total
            r += 1
        # total row
        _hdr(ws, r, 2, "Total", fill=TEAL)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=TEAL); ws.cell(r, 1).border = BORDER
        for i in range(8):
            _num(ws, r, 3 + i, round(tot[i], 2), bold=True, fill=TEAL)
        _num(ws, r, 11, round(tot[8], 2), bold=True, fill=TEAL)
        return r, tot

    entries = config["entries"]
    blockI = [e for e in entries if e.get("block", "I") == "I"]
    blockII = [e for e in entries if e.get("block") == "II"]

    r1, totI = write_block(5, "(I)  Classified under 9A / 9B / 9C", blockI, "I")
    r2, totII = write_block(r1 + 2, "(II)  Classified under 9D", blockII, "II")

    # grand total (I)+(II)
    gr = r2 + 1
    _hdr(ws, gr, 2, "G. Total (I)+(II)", fill=TEAL_DARK)
    ws.cell(gr, 1).fill = PatternFill("solid", fgColor=TEAL_DARK); ws.cell(gr, 1).border = BORDER
    grand = [round(totI[i] + totII[i], 2) for i in range(9)]
    for i in range(8):
        _num(ws, gr, 3 + i, grand[i], bold=True, fill=TEAL_DARK)
    _num(ws, gr, 11, grand[8], bold=True, fill=TEAL_DARK)

    widths = [5, 7, 11, 11, 11, 11, 10, 13, 11, 10, 13]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C7"

    # ---------------- GPVC sheet ----------------
    g = wb.create_sheet("GPVC")
    gh = ["SN", "Qtr", "Bill No", "Gross Amount", "Cement", "Reinforcement",
          "Other-section Steel", "Plate/Flat", "Balance for GPVC"]
    for c, h in enumerate(gh, 1):
        _hdr(g, 1, c, h)
    r = 2
    for e in entries:
        g.cell(r, 1, e["sn"]).border = BORDER
        g.cell(r, 2, e["qtr"]).border = BORDER
        g.cell(r, 3, e.get("bill", "")).border = BORDER
        gross = e.get("balance", 0) + e.get("reinf", 0) + e.get("other", 0) + e.get("cement", 0)
        for i, v in enumerate([gross, e.get("cement", 0), e.get("reinf", 0),
                               e.get("other", 0), e.get("plate", 0), e.get("balance", 0)]):
            _num(g, r, 4 + i, v)
        r += 1
    for c, w in enumerate([5, 7, 9, 15, 12, 14, 16, 12, 16], 1):
        g.column_dimensions[get_column_letter(c)].width = w
    g.freeze_panes = "D2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, grand[8]
